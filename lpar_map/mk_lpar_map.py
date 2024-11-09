from share.misc import save_json, load_json
import openpyxl
from openpyxl.utils import get_column_letter
from share.xl_styles import green, blue, grey, red, yellow
from collections import Counter
from make_table import get_lpar_table, get_sys_table, get_fcs_map_table, get_sn_table


def main():
    lpar_data = load_json('../lpar/files/lpar_data.json')
    lpar_table = get_lpar_table()
    sys_table = get_sys_table()
    fcs_map_table = get_fcs_map_table()
    sn_table = get_sn_table()
    make_xl(lpar_table, sys_table, fcs_map_table, sn_table)


def style_gen():
    while True:
        for style in [green, blue]:
            yield style


def mk_lpar_sheet(sheet, lpar_table):
    column_list = {'ufk_code': {'width': 15, 'head': 'Код региона'},
                   'ufk_name': {'width': 45, 'head': 'Название'},
                   'sys_name': {'width': 15, 'head': 'Power'},
                   'name': {'width': 18, 'head': 'Лпара'},
                   'ip': {'width': 15, 'head': 'ip-адрес'},
                   'os': {'width': 25, 'head': 'Версия OS'},
                   'desired_procs': {'width': 10, 'head': 'Des VP'},
                   'max_procs': {'width': 10, 'head': 'Max VP'},
                   'desired_proc_units': {'width': 10, 'head': 'Des PU'},
                   'max_proc_units': {'width': 10, 'head': 'Max PU'},
                   'desired_mem': {'width': 17, 'head': 'Des Memory (GB)'},
                   'max_mem': {'width': 20, 'head': 'Max Memory (GB)'},
                   'vg_name': {'width': 13, 'head': 'VG'},
                   'vg_size': {'width': 13, 'head': 'Размер (GB)'},
                   'vg_lun': {'width': 15, 'head': 'Число лунов'},
                   }
    sheet.title = 'Лпары'

    # Заголовки
    headers = [column_list[head]['head'] for head in column_list]
    sheet.append(headers)

    # Ширина столбцов
    for n, (column_name, column) in enumerate(column_list.items(), 1):
        sheet.column_dimensions[get_column_letter(n)].width = column['width']

    # Таблица
    for row in lpar_table:
        sheet.append(row)
    # Фиксация
    sheet.freeze_panes = 'A2'
    # Фильтры
    sheet.auto_filter.ref = "A1:D1"
    # Обьединение ячеек

    rows_count = Counter([x[0] for x in lpar_table])
    for row in sheet.rows:
        row_idx = row[0].row
        val = row[0].value
        count = rows_count[val]
        if rows_count[val] > 1:
            for col_idx in [1, 2]:
                sheet.merge_cells(start_row=row_idx, end_row=row_idx + count - 1, start_column=col_idx,
                                  end_column=col_idx)
    rows_count = Counter([x[3] for x in lpar_table])
    for row in sheet.rows:
        row_idx = row[3].row
        val = row[3].value
        count = rows_count[val]
        if rows_count[val] > 1 and val != 'vios1' and val != 'vios2':
            for col_idx in range(4, 13):
                sheet.merge_cells(start_row=row_idx, end_row=row_idx + count - 1, start_column=col_idx,
                                  end_column=col_idx)
    # Покраска
    style_generator = style_gen()
    style = next(style_generator)
    for row in sheet.rows:
        if row[0].row == 1:
            style = grey
        elif row[1].value == 'Не передан':
            style = red
        elif row[3].value == 'vios1' or row[3].value == 'vios2':
            style = yellow
        elif row[0].value:
            style = next(style_generator)
        if True in [True for x in row if x.value]:
            for cell in row:
                cell.style = style
    return sheet


def mk_sys_sheet(sheet, sys_table):
    columns = {
        'Сервер': 15,
        'S/N': 20,
        'ЦПУ всего': 15,
        'ЦПУ доступно': 20,
        'ЦПУ доступно с учётом выключенных': 37,
        'ОП всего, ГБ': 15,
        'ОП доступно, ГБ': 15,
        'ОП достпно с учётом выключенных, ГБ': 37,
        'Запущенно LPAR': 20,
        'HMC1 primary': 20,
        'HMC1 secondary': 20,
        'HMC2 primary': 20,
        'HMC2 secondary': 20,
    }
    # Заголовки
    headers = [column for column in columns]
    sheet.append(headers)

    # Ширина столбцов
    for n, (column_name, width) in enumerate(columns.items(), 1):
        sheet.column_dimensions[get_column_letter(n)].width = width
    # Таблица
    for row in sys_table:
        sheet.append(row)
    # Фиксация
    sheet.freeze_panes = 'B2'
    # Фильтры
    sheet.auto_filter.ref = "A1"
    # Покраска
    style_generator = style_gen()

    for row in sheet.rows:
        style = next(style_generator)
        if row[0].row == 1:
            style = grey
        for cell in row:
            cell.style = style
    return sheet


def mk_fcs_map_sheet(sheet, fcs_map_table):
    column_list = {
        'sys_name': {'width': 20, 'head': 'Лпара'},
        'name': {'width': 18, 'head': 'Power'},
        'fc': {'width': 15, 'head': 'FC адаптер'},
        'wwn': {'width': 25, 'head': 'WWN'},
        'vios': {'width': 20, 'head': 'VFC server'},
        'vios_fc': {'width': 15, 'head': 'VFC port'},
    }
    sheet.title = 'FC map'

    # Заголовки
    headers = [column_list[head]['head'] for head in column_list]
    sheet.append(headers)

    # Ширина столбцов
    for n, (column_name, column) in enumerate(column_list.items(), 1):
        sheet.column_dimensions[get_column_letter(n)].width = column['width']

    # Таблица
    for row in fcs_map_table:
        sheet.append(row)

    # Обьединение ячеек

    rows_count = Counter([x[0] for x in fcs_map_table])
    for row in sheet.rows:
        row_idx = row[0].row
        val = row[0].value
        count = rows_count[val]
        if rows_count[val] > 1:
            for col_idx in [1, 2]:
                sheet.merge_cells(start_row=row_idx, end_row=row_idx + count - 1, start_column=col_idx,
                                  end_column=col_idx)

    # Покраска
    style_generator = style_gen()
    style = next(style_generator)
    for row in sheet.rows:
        if row[0].row == 1:
            style = grey
        elif row[1].value == 'Не передан':
            style = red
        elif 'vios' in str(row[0].value):
            style = yellow
        elif row[0].value:
            style = next(style_generator)
        if True in [True for x in row if x.value]:
            for cell in row:
                cell.style = style
    return sheet


def mk_sn_sheet(sheet, sn_table):
    columns = {
        'Сервер': 15,
        'Серийный номер ПОСТГАРАНТ': 30,
    }
    # Заголовки
    headers = [column for column in columns]
    sheet.append(headers)

    # Ширина столбцов
    for n, (column_name, width) in enumerate(columns.items(), 1):
        sheet.column_dimensions[get_column_letter(n)].width = width

    # Таблица
    for row in sn_table:
        sheet.append(row)

    # Обьединение ячеек
    rows_count = Counter([x[0] for x in sn_table])
    for row in sheet.rows:
        row_idx = row[0].row
        val = row[0].value
        count = rows_count[val]
        if rows_count[val] > 1:
            sheet.merge_cells(start_row=row_idx, end_row=row_idx + count - 1, start_column=1,
                              end_column=1)
    # Покраска
    style_generator = style_gen()
    style = next(style_generator)
    for row in sheet.rows:
        if row[0].row == 1:
            style = grey
        elif row[0].value:
            style = next(style_generator)
        if True in [True for x in row if x.value]:
            for cell in row:
                cell.style = style


def make_xl(lpar_table, sys_table, fcs_map_table, sn_table):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet = mk_lpar_sheet(sheet, lpar_table)

    wb.create_sheet(title='Ресурсы Power')
    sheet = wb['Ресурсы Power']
    sheet = mk_sys_sheet(sheet, sys_table)

    wb.create_sheet(title='FC map')
    sheet = wb['FC map']
    sheet = mk_fcs_map_sheet(sheet, fcs_map_table)

    wb.create_sheet(title='S.N постгарант')
    sheet = wb['S.N постгарант']
    sheet = mk_sn_sheet(sheet, sn_table)

    wb.save('./export/GRD_LPARs_map.xlsx')


if __name__ == '__main__':
    main()
