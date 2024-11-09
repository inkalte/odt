from db import *
from share.host_db import HostDB
from share.misc import timeit, load_csv, save_json, load_json
from share.ssh import RemoteConnect, Output
from sqlalchemy.orm import Session, sessionmaker
from pprint import pprint
from dataclasses import dataclass
import openpyxl
from openpyxl.utils import get_column_letter
from share.xl_styles import green, blue, grey, red, yellow


@dataclass
class RmanData:
    host: str
    weekly_backup: str
    weekly_delete: str
    monthly_backup: str
    monthly_delete: str


def main():
    host_list = get_host_list()
    list_rman_data = get_list_rman_data(host_list)
    mk_xl(list_rman_data)


def mk_xl(list_rman_data: [RmanData]):
    wb = openpyxl.Workbook()
    wb.iso_dates = True
    sheet = wb.active
    columns = {
        'Сервер': 20,
        'weekly_backup': 100,
        'weekly_delete': 100,
        'monthly_backup': 100,
        'monthly_delete': 100,

    }
    # Заголовки
    headers = [column for column in columns]
    sheet.append(headers)

    # Ширина столбцов
    for n, (column_name, width) in enumerate(columns.items(), 1):
        sheet.column_dimensions[get_column_letter(n)].width = width
    # Таблица
    for row in list_rman_data:
        sheet.append([row.host, row.weekly_backup, row.weekly_delete, row.monthly_backup, row.monthly_delete])
    # Покраска
    style_generator = style_gen()

    for row in sheet.rows:
        style = next(style_generator)
        if row[0].row == 1:
            style = grey
        for cell in row:
            cell.style = style
    wb.save(f'./tmp/rman_script.xlsx')


def style_gen():
    while True:
        for style in [green, blue]:
            yield style


def get_list_rman_data(host_list: [Host]) -> [RmanData]:
    rman_data = []
    for host in host_list:
        print(host.hostname)
        rman_data.append(get_rman_data(host))
    return rman_data


def get_rman_data(host: Host) -> RmanData:
    rman_data = RmanData(
        host=host.hostname,
        weekly_backup='',
        weekly_delete='',
        monthly_backup='',
        monthly_delete='',
    )
    for script_type in ['WEEKLY', 'MONTHLY']:
        if host.type == "p04":
            tsm_path = "usr"
        else:
            tsm_path = "opt"
        cmd = f'grep -E "backup|delete" /{tsm_path}/tivoli/tsm/client/oracle/scripts/PROD_{script_type}.rman'
        ssh = RemoteConnect(host)
        output: Output = ssh.exec_command(cmd)
        output = output.stdout.read().decode().strip().split('\n')
        for row in output:
            if script_type == 'WEEKLY' and 'backup ' in row:
                rman_data.weekly_backup += row
            elif script_type == 'MONTHLY' and 'backup ' in row:
                rman_data.monthly_backup += row
            elif script_type == 'WEEKLY' and 'delete ' in row:
                rman_data.weekly_delete += row
            elif script_type == 'MONTHLY' and 'delete ' in row:
                rman_data.monthly_delete += row
    return rman_data


def get_host_list() -> [Host]:
    host_list = []
    for host in HostDB.iter_by_type('p04'):
        host_list.append(host)
    for host in HostDB.iter_by_type('r12'):
        host_list.append(host)
    return host_list


if __name__ == '__main__':
    main()
