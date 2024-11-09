import os
import csv
import stat
import time
from dataclasses import dataclass
from datetime import datetime
import pandas as pd
from db import Host
from share.host_db import HostDB
from share.misc import timeit, load_csv, save_json, load_json
from share.ssh import RemoteConnect, Output
import openpyxl
from openpyxl.comments import Comment
from share.xl_styles import green, blue, grey, red, yellow


def main():
    # update()
    events = load_json('tmp/events.json')
    node_list = []
    for n, event in enumerate(events):
        node_list.append(event['node_name'])
    node_list = sorted(set(node_list))

    node_list = {key: [] for key in node_list}
    for node in node_list:
        for event in events:
            if event['schedule_name'] in SCHEDS and event['node_name'] == node and event['status'] != 'Future':
                node_list[node].append(event)
    # for node in node_list:
    #     print(node, node_list[node])
    mk_xl(node_list)


def mk_xl(node_list: []):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.append(['NODE'])
    for row, (node, sheds) in enumerate(node_list.items(), 2):
        sheet.cell(row=row, column=1, value=node)
        m_10 = 2
        m_11 = 8
        m_12 = 14
        for col, shed in enumerate(sheds, 2):
            m = shed['scheduled_start'].split()[0].split('/')[0]
            if m == '10':
                col = m_10
                m_10 += 1
            elif m == '11':
                col = m_11
                m_11 += 1
            elif m == '12':
                col = m_12
                m_12 += 1
            else:
                continue

            cell = sheet.cell(row=row, column=col)
            cell.value = f"{shed['scheduled_start'].split()[0]}:{shed_type(shed['schedule_name'])}"
            if shed['status'] == 'Completed':
                cell.style = green
            elif shed['status'] == 'Failed':
                cell.style = red
            elif shed['status'] == 'Missed':
                cell.style = yellow
            # cell.comment = Comment(shed,'')
    wb.save('./tmp/ASFK_BIG_SHED.xlsx')


def shed_type(shed: str) -> str:
    if shed in ['SPRODWDBFULLM1700', 'APRODEDBFULLM1700', 'SPRODWDBFULLM0600', 'APRODEDBFULLM0100']:
        return 'Mon'
    elif shed in ['APRODEDBINC0W2300', 'SPRODWDBINC0W2300']:
        return 'Week'
    else:
        return shed


def update():
    events = []
    for tsm_node in HostDB.iter_by_type('tsm'):
        ssh = RemoteConnect(tsm_node)
        command = f'{DSMADMC} "q event * *00 n=*_ORACLE begind=-90 endd=today f=d"'
        output: Output = ssh.exec_command(command)
        with open(f'./tmp/{tsm_node.hostname}event.out', "wb") as file:
            file.write(output.stdout.read())

        with open(f'./tmp/{tsm_node.hostname}event.out', newline='') as csvfile:
            fieldnames = ['domain_name', 'schedule_name', 'node_name', 'scheduled_start', 'actual_start', 'completed',
                          'status',
                          'result', 'reason', ]
            reader = csv.DictReader(csvfile, delimiter='\t', fieldnames=fieldnames)
            for row in reader:
                events.append(row)
    save_json(events, 'tmp/events.json')


if __name__ == '__main__':
    DSMADMC = 'dsmadmc -id=zabbix -pass=123qweASD -TAB -DATAONLY=YES -outfile'
    SCHEDS = ['SPRODWDBFULLM1700', 'APRODEDBFULLM1700', 'SPRODWDBFULLM0600',
              'APRODEDBFULLM0100', 'APRODEDBINC0W2300', 'SPRODWDBINC0W2300']
    main()
