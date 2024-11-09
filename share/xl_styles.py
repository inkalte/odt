from openpyxl.styles import NamedStyle, Font, Border, Side, PatternFill, Alignment

bd = Side(style='thin', color="000000")
border = Border(left=bd, top=bd, right=bd, bottom=bd)
font = Font(name='Calibri', size=11)
alignment = Alignment(vertical='center', horizontal='center')

green = NamedStyle(name="green")
green.font = font
green.border = border
green.fill = PatternFill('solid', fgColor='c6e0b4')
green.alignment = alignment

blue = NamedStyle(name="blue")
blue.font = font
blue.border = border
blue.fill = PatternFill('solid', fgColor='b4c6e7')
blue.alignment = alignment

grey = NamedStyle(name="grey")
grey.font = font
grey.border = border
grey.fill = PatternFill('solid', fgColor='dbdbdb')
grey.alignment = alignment

yellow = NamedStyle(name="yellow")
yellow.font = font
yellow.border = border
yellow.fill = PatternFill('solid', fgColor='ffff66')
yellow.alignment = alignment

red = NamedStyle(name="red")
red.font = font
red.border = border
red.fill = PatternFill('solid', fgColor='ff9999')
red.alignment = alignment
