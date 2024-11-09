from share.misc import timeit, load_pickle
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from share.xl_styles import green, blue, grey, red, yellow
from srk_report import Event
from get_rman_logs import get_rman_logs
from get_errors import get_errors
from datetime import datetime


@timeit
def main():
    #errors = load_pickle('./tmp/parse_errors.pkl')
    errors = get_errors(begin_day='06/01/2024', end_day='today', test=False)
    errors = get_rman_logs(errors, event_filter=False, filter_str='')

    wb = openpyxl.Workbook()
    wb.iso_dates = True
    sheet = wb.active
    sheet = mk_sheet(sheet, get_table(errors))
    wb.save(f'C:/Users/AntipovDS/Desktop/ftp/srk_report_{datetime.now().strftime("%d.%m.%Y")}.xlsx')


def reason_translate(reason: str):
    if reason == 'shutdown':
        return 'Oracle недоступен'
    elif reason == 'ses_kill':
        return 'Отменён со стороны Oracle'
    elif reason == 'storage_space':
        return 'Недостаточно места'
    elif reason == 'connection_failure':
        return 'Ошибка подключения'
    elif reason == '9':
        return 'Отменён со стороны ОС (kill)'
    elif reason == 'none_logs':
        return f'код ошибки {reason}'
    else:
        return f'код ошибки {reason}'


def schedule_type_translate(shedule_type: str):
    if shedule_type == 'logs':
        return 'Логи'
    elif shedule_type == 'weekly':
        return 'Недельные'
    elif shedule_type == 'monthly':
        return 'Месячные'
    elif shedule_type == 'daily':
        return 'Суточные'
    else:
        return shedule_type


def get_table(errors):
    table = []
    for error in errors:
        if error.node_name.startswith('A'):
            info_sys = 'АСФК'
        elif error.node_name.startswith('S'):
            info_sys = 'СУФД'
        else:
            info_sys = ''
        error_text = f'Коллеги, добрый день, зафиксировали ошибку резервного копирования {info_sys} ' \
                     f'{schedule_type_translate(error.schedule_type)} ' \
                     f'{error.node_name} {error.completed.strftime("%d.%m.%Y %H:%M") if error.completed else error.actual_start.strftime("%d.%m.%Y %H:%M")}.' \
                     f' Подскажите пожалуйста, были остановки БД в этот период?'
        error: Event
        table.append([
            error.node_name,
            error.domain_name.replace('ASFK', '') + '00',
            error.completed,
            schedule_type_translate(error.schedule_type),
            reason_translate(error.reason),
            error_text,
            '',
            '',
            error.rman_log

        ])
    return table


def style_gen():
    while True:
        for style in [green, blue]:
            yield style


def mk_sheet(sheet, sys_table):
    columns = {
        'Сервер': 20,
        'УФК': 10,
        'Дата': 20,
        'Тип РК': 20,
        'Причина': 50,
        'Шаблон': 170,
        'Предупреждение': 20,
        'Ответ': 20,

    }
    # Заголовки
    headers = [column for column in columns]
    sheet.append(headers)

    # Ширина столбцов
    for n, (column_name, width) in enumerate(columns.items(), 1):
        sheet.column_dimensions[get_column_letter(n)].width = width
    # Таблица
    for row_idx, row in enumerate(sys_table, 2):
        sheet.append(row[0:7])
        if row[8]:
            sheet.cell(column=5, row=row_idx).comment = Comment(row[8], 'rman', width=600,
                                                                height=len(row[8].split('\n') * 25))
    # Покраска
    style_generator = style_gen()
    #
    for row in sheet.rows:
        style = next(style_generator)
        if row[0].row == 1:
            style = grey
        for cell in row:
            cell.style = style
            if cell.column == 3:
                cell.number_format = 'DD/MM/YYYY HH:MM'

    return sheet


if __name__ == '__main__':
    main()
